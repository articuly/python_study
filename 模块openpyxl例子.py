# coding:utf-8
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = 'python'
ws1 = wb.create_sheet('PHP')
print(wb.sheetnames)
ws['C1'] = '123'
ws.cell(row=2, column=4, value=456)
wb.save('python sample.xlsx')
